"""
Excel report generator for reconciliation results.
Creates multi-sheet workbooks with formatted output.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from ..models.transaction import (
    NormalizedTransaction,
    MatchResult,
    ReconciliationSummary,
)
from ..config import ReconConfig

logger = logging.getLogger(__name__)

# Style definitions
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
VARIANCE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
UNMATCHED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelReportGenerator:
    """Generates Excel reconciliation reports with multiple sheets."""

    def __init__(self, config: ReconConfig):
        """
        Initialize the report generator.

        Args:
            config: Application configuration
        """
        self.config = config
        output_config = config.output
        self.output_config = output_config.excel if hasattr(output_config, "excel") else {}
        self.sheet_config = output_config.sheets if hasattr(output_config, "sheets") else {}

    def generate_report(
        self,
        summary: ReconciliationSummary,
        matches: list[MatchResult],
        bank_only: list[NormalizedTransaction],
        intacct_only: list[NormalizedTransaction],
        output_path: Path,
    ) -> Path:
        """
        Generate the complete reconciliation report.

        Args:
            summary: Reconciliation summary
            matches: List of match results
            bank_only: Unmatched bank transactions
            intacct_only: Unmatched Intacct transactions
            output_path: Path for output file

        Returns:
            Path to generated report
        """
        logger.info(f"Generating Excel report: {output_path}")

        wb = Workbook()

        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        # Generate each sheet based on config
        summary_config = getattr(self.sheet_config, "summary", None)
        if summary_config is None or summary_config.enabled:
            self._create_summary_sheet(wb, summary)

        matched_config = getattr(self.sheet_config, "matched", None)
        if matched_config is None or matched_config.enabled:
            self._create_matched_sheet(wb, matches)

        bank_only_config = getattr(self.sheet_config, "bank_only", None)
        if bank_only_config is None or bank_only_config.enabled:
            self._create_bank_only_sheet(wb, bank_only)

        intacct_only_config = getattr(self.sheet_config, "intacct_only", None)
        if intacct_only_config is None or intacct_only_config.enabled:
            self._create_intacct_only_sheet(wb, intacct_only)

        variances_config = getattr(self.sheet_config, "amount_variances", None)
        if variances_config is None or variances_config.enabled:
            variance_matches = [m for m in matches if m.amount_variance]
            self._create_variance_sheet(wb, variance_matches)

        audit_config = getattr(self.sheet_config, "audit_trail", None)
        if audit_config is None or audit_config.enabled:
            self._create_audit_trail_sheet(wb, summary, matches)

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Report saved: {output_path}")

        return output_path

    def _create_summary_sheet(
        self, wb: Workbook, summary: ReconciliationSummary
    ) -> None:
        """Create the summary sheet with key metrics."""
        summary_config = getattr(self.sheet_config, "summary", None)
        sheet_name = summary_config.name if summary_config else "Summary"
        ws = wb.create_sheet(sheet_name)

        # Title
        ws["A1"] = "Bank Reconciliation Summary"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # File information section
        ws["A3"] = "File Information"
        ws["A3"].font = Font(bold=True)

        file_info = [
            ("BAI2 File:", summary.bai2_filename),
            ("Intacct File:", summary.intacct_filename),
            (
                "Reconciliation Date:",
                summary.reconciliation_date.strftime("%Y-%m-%d %H:%M:%S"),
            ),
            (
                "Statement Period:",
                f"{summary.statement_period_start} to {summary.statement_period_end}",
            ),
        ]

        for i, (label, value) in enumerate(file_info, start=4):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = str(value)

        # Transaction counts section
        ws["A9"] = "Transaction Counts"
        ws["A9"].font = Font(bold=True)

        count_data = [
            ("Total Bank Transactions:", summary.total_bank_transactions),
            ("Total Intacct Transactions:", summary.total_intacct_transactions),
            ("Matched Transactions:", summary.matched_count),
            ("Bank Only (Unmatched):", summary.bank_only_count),
            ("Intacct Only (Unmatched):", summary.intacct_only_count),
            ("Amount Variances:", summary.variance_count),
        ]

        for i, (label, value) in enumerate(count_data, start=10):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value

        # Match rates
        ws["A17"] = "Match Rates"
        ws["A17"].font = Font(bold=True)

        ws["A18"] = "Bank Match Rate:"
        ws["B18"] = f"{summary.match_rate_bank:.1f}%"
        ws["A19"] = "Intacct Match Rate:"
        ws["B19"] = f"{summary.match_rate_intacct:.1f}%"

        # Amount totals
        ws["A21"] = "Amount Totals"
        ws["A21"].font = Font(bold=True)

        amount_data = [
            ("Bank Total Credits:", f"${summary.bank_total_credits:,.2f}"),
            ("Bank Total Debits:", f"${summary.bank_total_debits:,.2f}"),
            ("Intacct Total Credits:", f"${summary.intacct_total_credits:,.2f}"),
            ("Intacct Total Debits:", f"${summary.intacct_total_debits:,.2f}"),
            ("Total Amount Variance:", f"${summary.total_amount_variance:,.2f}"),
        ]

        for i, (label, value) in enumerate(amount_data, start=22):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value

        # Matches by tier
        ws["A28"] = "Matches by Tier"
        ws["A28"].font = Font(bold=True)

        row = 29
        for tier, count in summary.matches_by_tier.items():
            ws[f"A{row}"] = tier
            ws[f"B{row}"] = count
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

    def _create_matched_sheet(self, wb: Workbook, matches: list[MatchResult]) -> None:
        """Create the matched transactions sheet."""
        matched_config = getattr(self.sheet_config, "matched", None)
        sheet_name = matched_config.name if matched_config else "Matched Transactions"
        ws = wb.create_sheet(sheet_name)

        headers = [
            "Bank Date",
            "Bank Reference",
            "Bank Amount",
            "Bank Description",
            "Intacct Date",
            "Intacct Reference",
            "Intacct Amount",
            "Intacct Description",
            "Match Tier",
            "Match Score",
            "Match Reason",
            "Amount Variance",
            "Date Variance (Days)",
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Write data
        for row_num, match in enumerate(matches, start=2):
            bank_txn = match.bank_transaction
            intacct_txn = (
                match.intacct_transactions[0] if match.intacct_transactions else None
            )

            row_data = [
                bank_txn.date,
                bank_txn.reference or "",
                float(bank_txn.amount),
                bank_txn.description,
                intacct_txn.date if intacct_txn else "",
                intacct_txn.reference if intacct_txn else "",
                float(intacct_txn.amount) if intacct_txn else "",
                intacct_txn.description if intacct_txn else "",
                match.match_tier,
                f"{match.match_score:.2f}",
                match.match_reason,
                float(match.amount_variance) if match.amount_variance else "",
                match.date_variance_days if match.date_variance_days else "",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER

                # Highlight variances
                if match.amount_variance and col in [12, 13]:
                    cell.fill = VARIANCE_FILL
                elif match.is_exact_match:
                    cell.fill = MATCH_FILL

        # Auto-fit columns
        self._auto_fit_columns(ws)

    def _create_bank_only_sheet(
        self, wb: Workbook, bank_only: list[NormalizedTransaction]
    ) -> None:
        """Create the bank-only transactions sheet."""
        bank_only_config = getattr(self.sheet_config, "bank_only", None)
        sheet_name = bank_only_config.name if bank_only_config else "Bank Only"
        ws = wb.create_sheet(sheet_name)

        headers = [
            "Date",
            "Reference",
            "Amount",
            "Type",
            "Description",
            "BAI2 Type Code",
            "BAI2 Type Description",
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Write data
        for row_num, txn in enumerate(bank_only, start=2):
            row_data = [
                txn.date,
                txn.reference or "",
                float(txn.amount),
                txn.type.value,
                txn.description,
                txn.bai2_type_code or "",
                txn.bai2_type_description or "",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER
                cell.fill = UNMATCHED_FILL

        self._auto_fit_columns(ws)

    def _create_intacct_only_sheet(
        self, wb: Workbook, intacct_only: list[NormalizedTransaction]
    ) -> None:
        """Create the Intacct-only transactions sheet."""
        intacct_only_config = getattr(self.sheet_config, "intacct_only", None)
        sheet_name = intacct_only_config.name if intacct_only_config else "Intacct Only"
        ws = wb.create_sheet(sheet_name)

        headers = [
            "Date",
            "Reference",
            "Amount",
            "Type",
            "Description",
            "Vendor",
            "GL Account",
            "Transaction ID",
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Write data
        for row_num, txn in enumerate(intacct_only, start=2):
            row_data = [
                txn.date,
                txn.reference or "",
                float(txn.amount),
                txn.type.value,
                txn.description,
                txn.vendor or "",
                txn.gl_account or "",
                txn.intacct_transaction_id or "",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER
                cell.fill = UNMATCHED_FILL

        self._auto_fit_columns(ws)

    def _create_variance_sheet(
        self, wb: Workbook, variance_matches: list[MatchResult]
    ) -> None:
        """Create the amount variances sheet."""
        variances_config = getattr(self.sheet_config, "amount_variances", None)
        sheet_name = variances_config.name if variances_config else "Amount Variances"
        ws = wb.create_sheet(sheet_name)

        headers = [
            "Reference",
            "Bank Date",
            "Bank Amount",
            "Intacct Date",
            "Intacct Amount",
            "Amount Variance",
            "Variance %",
            "Match Tier",
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Write data
        for row_num, match in enumerate(variance_matches, start=2):
            bank_txn = match.bank_transaction
            intacct_txn = (
                match.intacct_transactions[0] if match.intacct_transactions else None
            )

            variance_pct = ""
            if bank_txn.amount and match.amount_variance:
                variance_pct = f"{(match.amount_variance / bank_txn.amount) * 100:.2f}%"

            row_data = [
                bank_txn.reference or "",
                bank_txn.date,
                float(bank_txn.amount),
                intacct_txn.date if intacct_txn else "",
                float(intacct_txn.amount) if intacct_txn else "",
                float(match.amount_variance) if match.amount_variance else "",
                variance_pct,
                match.match_tier,
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER
                cell.fill = VARIANCE_FILL

        self._auto_fit_columns(ws)

    def _create_audit_trail_sheet(
        self, wb: Workbook, summary: ReconciliationSummary, matches: list[MatchResult]
    ) -> None:
        """Create the audit trail sheet."""
        audit_config = getattr(self.sheet_config, "audit_trail", None)
        sheet_name = audit_config.name if audit_config else "Audit Trail"
        ws = wb.create_sheet(sheet_name)

        # Reconciliation metadata
        ws["A1"] = "Reconciliation Audit Trail"
        ws["A1"].font = Font(size=14, bold=True)

        audit_info = [
            ("Generated At:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Config File:", summary.config_file_used or "Default"),
            ("Processing Time:", f"{summary.processing_time_seconds:.2f} seconds"),
            ("", ""),
            ("Match Statistics:", ""),
        ]

        row = 3
        for label, value in audit_info:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        # Match details by tier
        for tier, count in summary.matches_by_tier.items():
            ws[f"A{row}"] = f"  {tier}:"
            ws[f"B{row}"] = count
            row += 1

        row += 1
        ws[f"A{row}"] = "Match Log"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        # Headers for match log
        headers = ["Timestamp", "Bank ID", "Intacct ID", "Tier", "Score", "Reason"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row += 1

        # Match log entries
        for match in matches:
            intacct_ids = ", ".join(t.id for t in match.intacct_transactions)
            log_data = [
                match.matched_at.strftime("%Y-%m-%d %H:%M:%S"),
                match.bank_transaction.id,
                intacct_ids,
                match.match_tier,
                f"{match.match_score:.2f}",
                match.match_reason,
            ]

            for col, value in enumerate(log_data, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

        self._auto_fit_columns(ws)

    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """Auto-fit column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
